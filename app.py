import tkinter as tk
from tkinter import Menu, messagebox
from openpyxl import Workbook, load_workbook
from operacion import info  # Importar el diccionario desde operaciones.py

# Crear la ventana principal
app = tk.Tk()
app.title("Aplicación de Escritorio")
app.geometry("400x300")  # Tamaño de la ventana

# Crear un frame para el formulario y ocultarlo al inicio
formulario_pc_frame = tk.Frame(app)
formulario_pc_frame.pack(pady=20)

# Etiqueta y campo de entrada para "Código BP" dentro del frame
tk.Label(formulario_pc_frame, text="Código BP:").grid(row=0, column=0, padx=5, pady=5)
codigo_bp_entry = tk.Entry(formulario_pc_frame)
codigo_bp_entry.grid(row=0, column=1, padx=5, pady=5)

# Etiqueta y campo de entrada para "Dependencia" dentro del frame
tk.Label(formulario_pc_frame, text="Dependencia:").grid(row=1, column=0, padx=5, pady=5)
dependencia_entry = tk.Entry(formulario_pc_frame)
dependencia_entry.grid(row=1, column=1, padx=5, pady=5)

# Función para manejar el envío del formulario y guardarlo en Excel
def enviar_formulario():
    codigo_bp = codigo_bp_entry.get()
    dependencia = dependencia_entry.get()
    
    if codigo_bp and dependencia:
        # Crear o abrir el archivo de Excel
        try:
            # Si el archivo no existe, lo creamos
            try:
                wb = load_workbook("formulario_pc.xlsx")
                ws = wb.active
            except:
                wb = Workbook()
                ws = wb.active
                ws.title = "Formulario PC"
                # Escribir los encabezados si es la primera vez
                ws.append(["Código BP", "Marca", "Procesador", "Tarjeta Madre", "Memoria Ram", "Tarjeta de video", "Sistema Operativo", "Office", "Dependencia"])

            # Escribir los datos del formulario junto con los datos del diccionario
            datos = [codigo_bp] + list(info.values()) + [dependencia]
            ws.append(datos)

            # Guardar el archivo Excel
            wb.save("formulario_pc.xlsx")
            messagebox.showinfo("Formulario PC", "Datos guardados en Excel")

            # Limpiar los campos después de enviar
            codigo_bp_entry.delete(0, tk.END)
            dependencia_entry.delete(0, tk.END)
        except Exception as e:
            messagebox.showerror("Error", f"Hubo un error al guardar los datos en Excel: {str(e)}")
    else:
        messagebox.showwarning("Formulario PC", "Por favor, complete todos los campos.")

# Botón para enviar el formulario
tk.Button(formulario_pc_frame, text="Enviar", command=enviar_formulario).grid(row=2, column=0, columnspan=2, pady=10)

# Ocultar el frame inicialmente
formulario_pc_frame.pack_forget()

# Función para mostrar el formulario de "PC" dentro del frame
def mostrar_formulario_pc():
    # Mostrar el frame del formulario
    formulario_pc_frame.pack()

# Función para mostrar mensaje de "Impresora"
def mostrar_impresora():
    formulario_pc_frame.pack_forget()  # Ocultar el frame del formulario
    messagebox.showinfo("Impresora", "Has seleccionado la opción Impresora")

# Crear el menú de la aplicación
menu_bar = Menu(app)

# Crear el menú de opciones
opciones_menu = Menu(menu_bar, tearoff=0)
opciones_menu.add_command(label="PC", command=mostrar_formulario_pc)
opciones_menu.add_command(label="Impresora", command=mostrar_impresora)

# Agregar el menú de opciones a la barra de menú
menu_bar.add_cascade(label="Opciones", menu=opciones_menu)

# Configurar la barra de menú en la ventana
app.config(menu=menu_bar)

# Ejecutar la aplicación
app.mainloop()
